from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
#POR FAVOR, LEIA AS ORIENTAÇÕES EM READ-ME
#carregar planilha
caminho_arquivo_excel = r'CAMINHO PARA PLANILHA TESTE'
#Carregar webdriver
service = Service(r'CAMINHO PARA SEU WEBDRIVER')
driver = webdriver.Chrome(service=service)
driver.get('https://web.whatsapp.com/')
time.sleep(10)
#Abrir planilha, navegar pelas células enquanto houver conteúdo
arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)
nome_planilha = "teste" #caso crie uma nova worksheet, consulte o readme.md
planilha = arquivo_excel[nome_planilha]

for linha in planilha.iter_rows(min_row=1, max_row=planilha.max_row, values_only=True):
    link_celula = linha[0]
    driver.execute_script("window.open();")
    driver.switch_to.window(driver.window_handles[-1])
    driver.get(link_celula)
    time.sleep(2)
    #interagir com o navegador clicando automaticamente nos botões
    elemento_action = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "action-button")))
    elemento_action.click()
    elemento_usar = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.LINK_TEXT, "usar o WhatsApp Web")))
    elemento_usar.click()
    campo_texto = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true']")))
    #enviar mensagem
    botao_enviar = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span')))
    botao_enviar.click()
    time.sleep(2)
    #alternar entre abas
    driver.switch_to.window(driver.window_handles[-2])
    driver.close()
    driver.switch_to.window(driver.window_handles[-1])
#encerrar software
arquivo_excel.close()
time.sleep(5)
driver.quit()
